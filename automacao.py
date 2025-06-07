from ingest import carregar_documentos, gerar_chunks, criar_vetores
from rag import gerar_resposta
import os
import pandas as pd
import time
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

perguntas = [
    "O que é jubilamento?"
]

modelos_llm = {
    "Qwen3-235b": "qwen/qwen3-235b-a22b:free",
}


print("Carregando documentos e criando vetores...")
documentos = carregar_documentos()
chunks = gerar_chunks(documentos)
db = criar_vetores(chunks)


resultados = []

for i, pergunta in enumerate(perguntas, 1):
    print(f"\n==> Pergunta {i}: {pergunta}")
    linha = {"Pergunta": pergunta}
    for nome_modelo, id_modelo in modelos_llm.items():
        try:
            start = time.time()
            resposta = gerar_resposta(pergunta, db, modelo=id_modelo)
            end = time.time()
            if not resposta or not isinstance(resposta, str):
                raise ValueError("Resposta vazia ou inválida")
        except Exception as e:
            resposta = f"[Erro: {e}]"
            end = time.time()
        print(f"{nome_modelo}: {resposta[:60]}...")
        linha[nome_modelo] = resposta
        linha[f"{nome_modelo} (s)"] = round(end - start, 2)
    resultados.append(linha)

# Salva o DataFrame em Excel
df = pd.DataFrame(resultados)
output_path = "resultados_modelos.xlsx"
df.to_excel(output_path, index=False)

# == FORMATAÇÃO VISUAL COM OPENPYXL ==
wb = load_workbook(output_path)
ws = wb.active

# Estilo do cabeçalho
header_font = Font(bold=True)
fill_gray = PatternFill("solid", fgColor="DDDDDD")
alignment_wrap = Alignment(wrap_text=True, vertical="top")

# Formata o cabeçalho
for cell in ws[1]:
    cell.font = header_font
    cell.fill = fill_gray
    cell.alignment = alignment_wrap

# Cores alternadas nas linhas
fill_alt = PatternFill("solid", fgColor="F5F5F5")
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    if row[0].row % 2 == 0:
        for cell in row:
            cell.fill = fill_alt
            cell.alignment = alignment_wrap
    else:
        for cell in row:
            cell.alignment = alignment_wrap

# Ajuste automático de largura das colunas
for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    adjusted_width = min(max_length + 5, 80)
    ws.column_dimensions[col_letter].width = adjusted_width

# Salva a planilha formatada
wb.save(output_path)

print(f"\nTodas as respostas foram salvas em: {output_path}")
