from ingest import carregar_documentos, gerar_chunks, criar_vetores
from rag import gerar_resposta
import os
import pandas as pd
import time
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

perguntas = [
    "Qual documento é necessário para compensação de falta por motivo religioso?",
    "Quando pode ter dispensa de aula?",
    "O que é jubilamento?",
    "Onde devo fazer o pedido de compensação de falta por motivo religioso?",
    "O que é a compensação de falta por motivo religioso?",
    "O que são atividades acompanhadas?",
    "Quem pode solicitar atividades acompanhadas?",
    "Onde eu devo fazer o pedido de atividades acompanhadas?",
    "Tenho prazo para fazer o pedido de atividades acompanhadas?",
    "Quem tem direito a compensação de falta?",
    "O que é a compensação de falta?",
    "Posso reingressar no curso?",
    "Quando eu posso ser desligado do curso?",
    "O que é o exame de suficiência?",
    "Posso pedir revisão de prova?",
    "Quantas vezes posso pedir o exame de suficiência?",
    "Perdi prova por motivo de doença, o que devo fazer?",
    "Qual o critério para ser aprovado em uma disciplina?",
    "Quantas disciplinas de enriquecimento curricular eu posso me matricular?",
    "O que são disciplinas de enriquecimento curricular?",
    "Posso me matricular em disciplinas que ocorre no mesmo dia e horário?",
    "O que são turmas sem presença obrigatória?",
    "Quem pode fazer matrícula nas turma sem presença obrigatória?",
    "Qual é quantidade máxima de disciplinas que eu posso me matricular?",
    "Como faço para quebrar o pré-requisito de uma disciplina que tenho dependência?",
    "O que acontece se eu esquecer de fazer a matrícula?",
    "Como eu calculo qual é o período em que eu estou?",
    "Como é calculado o coeficiente absoluto?",
    "Como é calculado o coeficiente normalizado?",
    "O que significa ficar dependente em uma disciplina?",
    "Ao solicitar a matrícula, eu tenho vaga garantida?",
    "Posso me matricular em disciplinas de outros períodos do meu curso?",
    "Qual é a ordem de preenchimento das vagas das disciplinas?",
    "O que é aluno formando?",
    "O que é convalidação de disciplina?",
    "Quais disciplinas podem ser convalidadas?",
    "Quando pode ser interessante fazer matrícula em turmas sem presença obrigatória?",
    "Quantas disciplinas eu posso me matricular em turma sem presença obrigatória?",
    "Quem faz a convalidação de disciplinas?",
    "Quais os critérios usados para convalidar ou não uma disciplina?",
    "Posso usar mais de uma disciplina cursada para convalidar uma disciplina?",
    "Minha nota na disciplina é via conceito. O que é feito para convalidar nesse caso?",
    "No meu histórico escolar não tem porcentagem de frequência. O que acontece nesse caso para convalidar?",
    "Posso cursar disciplinas de outros cursos?",
    "Como me matriculo em disciplinas de outros cursos?",
    "Qual é o número mínimo de avaliações de uma disciplina?",
    "Posso pedir exame de suficiência de qualquer disciplina?",
    "Posso pedir o exame de suficiência sem estar matriculado na disciplina?",
    "O que são turmas intensivas?",
    "O que é turma de férias?",
    "O que é o trancamento do curso?",
    "Posso trancar no primeiro período?",
    "O que é o cancelamento de disciplina?",
    "Posso cancelar disciplinas?",
    "Como eu cancelo uma disciplina?",
    "Qual o prazo máximo para conclusão de curso?",
    "O que é risco de jubilamento?",
    "O que é o desligamento do curso?",
    "Preciso adicionar algum documento para pedir atividades acompanhadas?",
    "Qual o tempo de duração das atividades acompanhadas?",
    "Quando posso ter abono de falta?",
    "Como eu peço compensação de falta?",
    "Quantas horas eu devo cumprir para ser aprovado na disciplina de estágio em engenharia de software?",
    "Quais tipos de estágio existem no curso de engenharia de software?",
    "Em quais semestres eu posso fazer o estágio obrigatório em engenharia de software?",
    "Posso fazer o estágio obrigatório em mais de uma empresa?",
    "Quais relatórios devo entregar se eu fizer os dois tipos de estágio obrigatório (convencional e validação)?",
    "A declaração da organização de estágio deve ter a assinatura da empresa mesmo (pessoa jurídica) ou pessoa física?",
    "Projeto de ensino, pesquisa ou extensão podem ser utilizados para validar estágio obrigatório?"
]

modelos_llm = {
    'GPT 4o': 'gpt-4o-mini',
    'DeepSeek R1': 'deepseek/deepseek-r1-0528-qwen3-8b:free',
    'LLama 4 Scout': 'meta-llama/llama-4-scout:free',
    'Gemini 2.0 Flash': 'google/gemini-2.0-flash-exp:free',
    "Gemma 3n": 'google/gemma-3n-e4b-it:free',
    "Phy 4": "microsoft/phi-4-reasoning:free",
    "Qwen3-235b": "qwen/qwen3-235b-a22b:free",
}


print("Carregando documentos e criando vetores...")
documentos = carregar_documentos()
chunks = gerar_chunks(documentos)
db = criar_vetores(chunks)


resultados = []

for i, pergunta in enumerate(perguntas, 1):
    print(f"\n==> Pergunta {i}: {pergunta}")
    linha = {"Pergunta": pergunta}
    for nome_modelo, id_modelo in modelos_llm.items():
        try:
            start = time.time()
            resposta = gerar_resposta(pergunta, db, modelo=id_modelo)
            end = time.time()
            if not resposta or not isinstance(resposta, str):
                raise ValueError("Resposta vazia ou inválida")
        except Exception as e:
            resposta = f"[Erro: {e}]"
            end = time.time()
        print(f"{nome_modelo}: {resposta[:60]}...")
        linha[nome_modelo] = resposta
        linha[f"{nome_modelo} (s)"] = round(end - start, 2)
    resultados.append(linha)

# Salva o DataFrame em Excel
df = pd.DataFrame(resultados)
output_path = "resultados_modelos.xlsx"
df.to_excel(output_path, index=False)

# == FORMATAÇÃO VISUAL COM OPENPYXL ==
wb = load_workbook(output_path)
ws = wb.active

# Estilo do cabeçalho
header_font = Font(bold=True)
fill_gray = PatternFill("solid", fgColor="DDDDDD")
alignment_wrap = Alignment(wrap_text=True, vertical="top")

# Formata o cabeçalho
for cell in ws[1]:
    cell.font = header_font
    cell.fill = fill_gray
    cell.alignment = alignment_wrap

# Cores alternadas nas linhas
fill_alt = PatternFill("solid", fgColor="F5F5F5")
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    if row[0].row % 2 == 0:
        for cell in row:
            cell.fill = fill_alt
            cell.alignment = alignment_wrap
    else:
        for cell in row:
            cell.alignment = alignment_wrap

# Ajuste automático de largura das colunas
for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    adjusted_width = min(max_length + 5, 80)
    ws.column_dimensions[col_letter].width = adjusted_width

# Salva a planilha formatada
wb.save(output_path)

print(f"\nTodas as respostas foram salvas em: {output_path}")
